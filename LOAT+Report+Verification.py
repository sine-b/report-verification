
# coding: utf-8

# In[ ]:

import pandas as pd
import numpy as np
from pandas import DataFrame
from collections import defaultdict

import openpyxl


# In[ ]:

apList = pd.read_excel('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/APPENDIX B - IFM Global List of APs.xlsx')
traderList = pd.read_excel('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/APPENDIX C - Authorized Traders and LPTs.xlsx')

wb = openpyxl.load_workbook('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/excel loat template.xlsx')
ws = wb.worksheets[0]

sortedAP = apList.sort_values(['AP Name'])
apList=apList.reindex()
boolean = False
traderBool = False


# In[ ]:

def findAPName(sortedAP):
    nameDict = {}
    global boolean
    for index, row in sortedAP.iterrows():
        AP = row['AP Name']
        if AP[-1] == " ":
            AP = AP[:-1]
        splitAP = AP.split(',')
        newName = splitAP[1] + " " + splitAP[0]
        newName = newName[1:]
        
        city = row['City']
        splitCity = city.split(',')
        newCity = splitCity[0]
        nameDict.setdefault(newCity,[]).append(newName.lower())
    
    while not boolean:
        boolo = False
        nameInput = raw_input("Enter Broker Name (Q to quit): ")
        if nameInput.lower() == "q":
            break
        while not boolo:
            for search in nameDict.values():
                if nameInput.lower() in search:
                    print "Broker is an AP"
                    boolean = True
                    return (nameInput, nameDict)
            else:
                print "Broker is not an AP"
                break


# In[ ]:

def applyAPName(excelLoat, nameInput, nameDict):
    
    ws['C8'] = nameInput.title()
    for search in nameDict.values():
        if nameInput.lower() in search:
            ws['D8'] = "(" + nameDict.keys()[nameDict.values().index(search)] + ")"
    #wb.save('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/modExcelLoat.xlsx')


# In[ ]:

def findTraderName(traderList):
    trader = ""
    city = ""
    nameDict = {}
    global traderBool
    for index, row in traderList.iterrows():
        if (row['Trader'] is "AUTHORIZED TRADERS") or (type(row['Trader']) is float) or (type(row['Location']) is float):
            pass
        else:
            trader = row['Trader']
        if row['Location'] is "LIMITED PURPOSE TRADERS":
            pass
        if type(row['Location']) is float:
            pass
        else:
            city = row['Location']
        #print type(city)
        if "Chicago" in city:
            city = "chicago"
        elif "Kansas City" in city:
            city = "kansas city"
        elif "New York" in city:
            city = "new york"
        elif "Miami" in city:
            city = "miami"
        elif "Asuncion" in city:
            city = "asuncion"
        elif "Bloomington" in city:
            city = "bloomington"
        elif "Campinas" in city:
            city = "campinas"
        elif "Sao Paolo" in city:
            city = "sao paolo"
        elif "Sydney" in city:
            city = "sydney"
        elif "West Des Moines" in city:
            city = "west des moines"
        elif "Bowling Green" in city:
            city = "bowling green"
        elif "Buenos Aires" in city:
            city = "bueno aires"
        nameDict.setdefault(city,[]).append(trader.lower())
        
    while not traderBool:
        boole = False
        nameInput = raw_input("Enter Trader/LPT Name (Q to quit): ")
        if nameInput.lower() == "q":
            break
        while not boole:
            for search in nameDict.values():
                if nameInput.lower() in search:
                    print "Trader/LPT is eligible"
                    boolean = True
                    return (nameInput, nameDict)
            else:
                print "Trader/LPT is not eligible"
                break


# In[ ]:

def applyTraderName(excelLoat, nameInput, nameDict):
    ws['C9'] = nameInput.title()
    for search in nameDict.values():
        if nameInput.lower() in search:
            ws['D9'] = "(" + (nameDict.keys()[nameDict.values().index(search)]).title() + ")"
    #wb.save('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/modExcelLoat.xlsx')


# In[ ]:

def traderStatus(traderList, excelLoat, traderName):
    aList = []
    refIndex = 0
    for trader in traderList['Trader']:
        if pd.isnull(trader) is False:
            aList.append(trader.lower())
    for index, item in enumerate(aList):
        if item == 'LIMITED PURPOSE TRADERS'.lower():
            refIndex = index
    if aList.index(traderName) < refIndex:
        ws['C10'] = "Authorized Trader"
    else:
        ws['C10'] = "Limited Purpose Trader"
    #wb.save('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/modExcelLoat.xlsx')


# In[ ]:

def miscInfo(excelLoat):
    repDate = raw_input("Date of report: ")
    repNum = input("LOAT report number: ")
    trdDate = raw_input("Trade date (MM/DD/YYYY): ")
    acctNum = input("Customer acct number: ")
    trdID = input("Trade ID: ")
    lang = raw_input("Languages used: ")
    dtccStat = raw_input("DTCC status: ")
    
    ws['A3'] = ws['A3'].value + repDate
    ws['C7'] = repNum
    ws['C11'] = trdDate
    ws['C11'].number_format = 'mm/dd/yyyy'
    ws['C12'] = acctNum
    ws['C13'] = trdID
    ws['C14'] = lang.title()
    ws['C15'] = dtccStat.title()
    #wb.save('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/modExcelLoat.xlsx')


# In[ ]:

def runAll():
    apNameOutput = findAPName(sortedAP)
    if apNameOutput is not None:
        applyAPName(ws, apNameOutput[0], apNameOutput[1])
    traderNameOutput = findTraderName(traderList)
    if traderNameOutput is not None:
        applyTraderName(ws, traderNameOutput[0], traderNameOutput[1])
    traderStatus(traderList, ws, traderNameOutput[0])
    miscInfo(ws)
    wb.save('C:/Users/ethan.dam.FCSTONE/Downloads/LOAT proj/modExcelLoat.xlsx')


# In[ ]:

runAll()

